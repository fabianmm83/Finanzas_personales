import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
import os
import io
from datetime import datetime

# Configuración
DB_PATH = 'instance/app.db'
REPORT_FILE = 'salud_financiera_mensual.xlsx'
IMAGE_FOLDER = 'temp_charts'
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# Colores para el Excel
HEADER_COLOR = '002060'
POSITIVE_COLOR = 'C6EFCE'
NEGATIVE_COLOR = 'FFC7CE'

# Conectar a la base de datos
def connect_db():
    return sqlite3.connect(DB_PATH)

# Obtener datos de transacciones
def get_transactions_data(conn):
    query = """
    SELECT date, description, amount, category, type
    FROM transactions
    ORDER BY date
    """
    df = pd.read_sql_query(query, conn)
    
    # Procesamiento de fechas
    df['date'] = pd.to_datetime(df['date'])
    df['mes'] = df['date'].dt.to_period('M')
    df['mes_str'] = df['mes'].astype(str)
    df['año'] = df['date'].dt.year
    df['mes_num'] = df['date'].dt.month
    
    # DEBUG: Verificar los datos originales
    print("Datos originales:")
    print(f"Total transacciones: {len(df)}")
    print(f"Distribución por tipo:")
    print(df['type'].value_counts())
    print(f"Rango de montos: {df['amount'].min()} to {df['amount'].max()}")
    
    return df

# Análisis mensual básico
def basic_monthly_analysis(df):
    # Crear DataFrames separados para ingresos y gastos basados en el campo 'type'
    ingresos_df = df[df['type'] == 'i'].copy()
    gastos_df = df[df['type'] == 'g'].copy()
    
    print(f"\nIngresos encontrados: {len(ingresos_df)} transacciones")
    print(f"Gastos encontrados: {len(gastos_df)} transacciones")
    
    # Agrupar ingresos por mes
    if not ingresos_df.empty:
        ingresos_mensual = ingresos_df.groupby(['año', 'mes_num', 'mes_str']).agg(
            ingresos=('amount', 'sum'),
            count_ingresos=('amount', 'count')
        ).reset_index()
    else:
        ingresos_mensual = pd.DataFrame(columns=['año', 'mes_num', 'mes_str', 'ingresos', 'count_ingresos'])
    
    # Agrupar gastos por mes
    if not gastos_df.empty:
        gastos_mensual = gastos_df.groupby(['año', 'mes_num', 'mes_str']).agg(
            gastos=('amount', 'sum'),
            count_gastos=('amount', 'count')
        ).reset_index()
    else:
        gastos_mensual = pd.DataFrame(columns=['año', 'mes_num', 'mes_str', 'gastos', 'count_gastos'])
    
    # Obtener todos los meses únicos
    all_months = df[['año', 'mes_num', 'mes_str']].drop_duplicates()
    
    # Hacer merge completo
    merged = pd.merge(all_months, ingresos_mensual, on=['año', 'mes_num', 'mes_str'], how='left')
    merged = pd.merge(merged, gastos_mensual, on=['año', 'mes_num', 'mes_str'], how='left')
    
    # Llenar valores NaN con 0
    merged['ingresos'] = merged['ingresos'].fillna(0)
    merged['gastos'] = merged['gastos'].fillna(0)
    merged['count_ingresos'] = merged['count_ingresos'].fillna(0)
    merged['count_gastos'] = merged['count_gastos'].fillna(0)
    
    # Calcular métricas adicionales
    merged['ahorro'] = merged['ingresos'] - merged['gastos']
    merged['ratio_gastos'] = merged.apply(
        lambda x: x['gastos'] / x['ingresos'] if x['ingresos'] > 0 else 0, 
        axis=1
    ).round(4)
    
    # Ordenar por año y mes
    merged = merged.sort_values(['año', 'mes_num'])
    
    print("\nResumen mensual calculado:")
    print(merged[['mes_str', 'ingresos', 'gastos', 'ahorro']].to_string())
    
    return merged, ingresos_df, gastos_df

# Análisis de categorías
def category_analysis(ingresos_df, gastos_df):
    # Gastos por categoría
    if not gastos_df.empty:
        gastos_cat = gastos_df.groupby(['category']).agg(
            total=('amount', 'sum'),
            count=('amount', 'count'),
            avg=('amount', 'mean')
        ).reset_index()
        gastos_cat = gastos_cat.sort_values('total', ascending=False)
        print(f"\nGastos por categoría: {len(gastos_cat)} categorías")
        print(gastos_cat.head())
    else:
        gastos_cat = pd.DataFrame(columns=['category', 'total', 'count', 'avg'])
        print("\nNo se encontraron gastos por categoría")
    
    # Ingresos por categoría
    if not ingresos_df.empty and 'category' in ingresos_df.columns:
        ingresos_cat = ingresos_df.groupby(['category']).agg(
            total=('amount', 'sum'),
            count=('amount', 'count'),
            avg=('amount', 'mean')
        ).reset_index()
        ingresos_cat = ingresos_cat.sort_values('total', ascending=False)
        print(f"Ingresos por categoría: {len(ingresos_cat)} categorías")
        print(ingresos_cat.head())
    else:
        ingresos_cat = None
        print("No se encontraron ingresos por categoría")
    
    return gastos_cat, ingresos_cat

# Crear gráficas
def create_charts(merged, gastos_cat):
    charts = {}
    
    # Gráfica de flujo mensual
    plt.figure(figsize=(12, 6))
    
    # Solo graficar si hay datos
    if not merged.empty and merged['ingresos'].sum() + merged['gastos'].sum() > 0:
        x_pos = range(len(merged['mes_str']))
        
        plt.bar(x_pos, merged['ingresos'], color='green', alpha=0.7, label='Ingresos', width=0.4)
        plt.bar([x + 0.4 for x in x_pos], merged['gastos'], color='red', alpha=0.7, label='Gastos', width=0.4)
        
        # Solo graficar ahorro si hay datos significativos
        if abs(merged['ahorro'].max() - merged['ahorro'].min()) > 0.01:
            plt.plot(x_pos, merged['ahorro'], marker='o', color='blue', linewidth=2, label='Ahorro')
        
        plt.title('Flujo Financiero Mensual')
        plt.xlabel('Mes')
        plt.ylabel('Monto ($)')
        plt.xticks(x_pos, merged['mes_str'], rotation=45)
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
    else:
        plt.text(0.5, 0.5, 'No hay datos suficientes\npara generar el gráfico', 
                ha='center', va='center', transform=plt.gca().transAxes)
        plt.title('Flujo Financiero Mensual - Sin Datos')
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    buf.seek(0)
    charts['flujo_mensual'] = buf
    plt.close()
    
    # Gráfica de ratio gastos/ingresos (solo si hay ingresos)
    plt.figure(figsize=(12, 6))
    if not merged.empty and merged['ingresos'].sum() > 0:
        x_pos = range(len(merged['mes_str']))
        plt.bar(x_pos, merged['ratio_gastos'], color='purple', alpha=0.7)
        plt.axhline(y=0.7, color='r', linestyle='--', label='Límite recomendado (70%)')
        plt.title('Ratio Gastos/Ingresos por Mes')
        plt.xlabel('Mes')
        plt.ylabel('Ratio (Gastos/Ingresos)')
        plt.xticks(x_pos, merged['mes_str'], rotation=45)
        plt.legend()
        plt.grid(True, alpha=0.3)
    else:
        plt.text(0.5, 0.5, 'No hay datos suficientes\npara generar el gráfico', 
                ha='center', va='center', transform=plt.gca().transAxes)
        plt.title('Ratio Gastos/Ingresos - Sin Datos')
    
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    buf.seek(0)
    charts['ratio_gastos'] = buf
    plt.close()
    
    # Gráfica de categorías de gastos
    plt.figure(figsize=(12, 6))
    if len(gastos_cat) > 0 and not gastos_cat.empty:
        top_gastos = gastos_cat.head(10)
        plt.barh(top_gastos['category'], top_gastos['total'], color='orange', alpha=0.7)
        plt.title('Top 10 Categorías de Gastos')
        plt.xlabel('Monto Total ($)')
    else:
        plt.text(0.5, 0.5, 'No hay datos de gastos\npor categoría', 
                ha='center', va='center', transform=plt.gca().transAxes)
        plt.title('Categorías de Gastos - Sin Datos')
    
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    buf.seek(0)
    charts['top_gastos'] = buf
    plt.close()
    
    return charts

# Crear reporte Excel
def create_excel_report(merged, gastos_cat, ingresos_cat, charts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Financiero"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    positive_fill = PatternFill(start_color=POSITIVE_COLOR, end_color=POSITIVE_COLOR, fill_type="solid")
    negative_fill = PatternFill(start_color=NEGATIVE_COLOR, end_color=NEGATIVE_COLOR, fill_type="solid")
    
    # Título
    ws['A1'] = "Análisis de Salud Financiera Mensual"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    # Resumen Mensual
    ws['A3'] = "Resumen Mensual"
    ws['A3'].font = Font(bold=True, size=14)
    
    headers = ['Mes', 'Ingresos', 'Gastos', 'Ahorro', 'Ratio Gastos/Ingresos']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, (_, data) in enumerate(merged.iterrows(), 5):
        ws.cell(row=row, column=1, value=data['mes_str'])
        
        # Ingresos
        ws.cell(row=row, column=2, value=data['ingresos'])
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        
        # Gastos
        gasto_cell = ws.cell(row=row, column=3, value=data['gastos'])
        gasto_cell.number_format = '"$"#,##0.00'
        
        # Ahorro
        ahorro_cell = ws.cell(row=row, column=4, value=data['ahorro'])
        ahorro_cell.number_format = '"$"#,##0.00'
        if data['ahorro'] > 0:
            ahorro_cell.fill = positive_fill
        else:
            ahorro_cell.fill = negative_fill
        
        # Ratio
        ratio_cell = ws.cell(row=row, column=5, value=data['ratio_gastos'])
        ratio_cell.number_format = '0.00%'
        if data['ratio_gastos'] > 0.7:
            ratio_cell.fill = negative_fill
    
    # Ajustar anchos de columnas
    for column in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[column].width = 18
    
    # Insights Automáticos
    current_row = len(merged) + 7
    
    # Solo calcular promedios si hay datos
    if not merged.empty:
        avg_ingresos = merged['ingresos'].mean()
        avg_gastos = merged['gastos'].mean()
        avg_ahorro = merged['ahorro'].mean()
        avg_ratio = merged['ratio_gastos'].mean()
        
        ws.cell(row=current_row, column=1, value="Promedios:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        ws.cell(row=current_row, column=2, value=avg_ingresos)
        ws.cell(row=current_row, column=2).number_format = '"$"#,##0.00'
        
        ws.cell(row=current_row, column=3, value=avg_gastos)
        ws.cell(row=current_row, column=3).number_format = '"$"#,##0.00'
        
        ws.cell(row=current_row, column=4, value=avg_ahorro)
        ws.cell(row=current_row, column=4).number_format = '"$"#,##0.00'
        
        ws.cell(row=current_row, column=5, value=avg_ratio)
        ws.cell(row=current_row, column=5).number_format = '0.00%'
        
        current_row += 2
        
        # Mejor y peor mes
        if len(merged) > 1:
            best_month = merged.loc[merged['ahorro'].idxmax()]
            worst_month = merged.loc[merged['ahorro'].idxmin()]
            
            ws.cell(row=current_row, column=1, value="Mejor Mes:")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=f"{best_month['mes_str']} (Ahorro: ${best_month['ahorro']:,.2f})")
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Peor Mes:")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=f"{worst_month['mes_str']} (Ahorro: ${worst_month['ahorro']:,.2f})")
            current_row += 2
        
        # Recomendaciones
        if avg_ratio > 0.7:
            ws.cell(row=current_row, column=1, value="ALERTA: Tu ratio promedio de gastos/ingresos es alto (>70%)")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000")
            current_row += 1
            ws.cell(row=current_row, column=1, value="Recomendación: Reduce gastos en categorías no esenciales")
            current_row += 2
        
        if avg_ahorro < 0:
            ws.cell(row=current_row, column=1, value="ALERTA: En promedio estás gastando más de lo que ingresas")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000")
            current_row += 1
            ws.cell(row=current_row, column=1, value="Recomendación: Revisa tus gastos fijos y considera aumentar ingresos")
            current_row += 2
    else:
        ws.cell(row=current_row, column=1, value="No hay datos suficientes para análisis")
        current_row += 2
    
    # Añadir gráficas
    try:
        img = Image(charts['flujo_mensual'])
        img.width = 600
        img.height = 300
        ws.add_image(img, f'A{current_row}')
        current_row += 20
    except:
        ws.cell(row=current_row, column=1, value="No se pudo generar gráfico de flujo mensual")
        current_row += 2
    
    try:
        img = Image(charts['ratio_gastos'])
        img.width = 600
        img.height = 300
        ws.add_image(img, f'A{current_row}')
        current_row += 20
    except:
        ws.cell(row=current_row, column=1, value="No se pudo generar gráfico de ratio")
        current_row += 2
    
    try:
        if 'top_gastos' in charts:
            img = Image(charts['top_gastos'])
            img.width = 600
            img.height = 300
            ws.add_image(img, f'A{current_row}')
            current_row += 20
    except:
        ws.cell(row=current_row, column=1, value="No se pudo generar gráfico de categorías")
        current_row += 2
    
    # Hoja de Gastos por Categoría
    if len(gastos_cat) > 0 and not gastos_cat.empty:
        ws_gastos = wb.create_sheet("Gastos por Categoría")
        
        ws_gastos.append(['Categoría', 'Total Gastado', 'N° Transacciones', 'Promedio por Transacción'])
        for _, row in gastos_cat.iterrows():
            ws_gastos.append([
                row['category'],
                row['total'],
                row['count'],
                row['avg']
            ])
        
        # Formato
        for cell in ws_gastos[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for row in ws_gastos.iter_rows(min_row=2, max_col=4, max_row=len(gastos_cat)+1):
            row[1].number_format = '"$"#,##0.00'
            row[3].number_format = '"$"#,##0.00'
        
        # Ajustar anchos
        ws_gastos.column_dimensions['A'].width = 25
        ws_gastos.column_dimensions['B'].width = 15
        ws_gastos.column_dimensions['C'].width = 15
        ws_gastos.column_dimensions['D'].width = 20
    
    # Hoja de Ingresos por Categoría (si aplica)
    if ingresos_cat is not None and len(ingresos_cat) > 0:
        ws_ingresos = wb.create_sheet("Ingresos por Categoría")
        
        ws_ingresos.append(['Categoría', 'Total Ingresado', 'N° Transacciones', 'Promedio por Transacción'])
        for _, row in ingresos_cat.iterrows():
            ws_ingresos.append([
                row['category'],
                row['total'],
                row['count'],
                row['avg']
            ])
        
        # Formato
        for cell in ws_ingresos[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for row in ws_ingresos.iter_rows(min_row=2, max_col=4, max_row=len(ingresos_cat)+1):
            row[1].number_format = '"$"#,##0.00'
            row[3].number_format = '"$"#,##0.00'
        
        # Ajustar anchos
        ws_ingresos.column_dimensions['A'].width = 25
        ws_ingresos.column_dimensions['B'].width = 15
        ws_ingresos.column_dimensions['C'].width = 15
        ws_ingresos.column_dimensions['D'].width = 20
    
    # Guardar el archivo
    wb.save(REPORT_FILE)
    print(f"Reporte Excel guardado como: {REPORT_FILE}")

# Función principal
def main():
    print("Iniciando análisis de salud financiera...")
    
    try:
        conn = connect_db()
        print("Conectado a la base de datos")
        
        # Obtener datos
        print("Obteniendo datos de transacciones...")
        df = get_transactions_data(conn)
        
        if df.empty:
            print("No se encontraron transacciones en la base de datos.")
            return
        
        print(f"Analizando {len(df)} transacciones...")
        
        # Análisis básico
        merged, ingresos_df, gastos_df = basic_monthly_analysis(df)
        
        # Análisis por categoría
        gastos_cat, ingresos_cat = category_analysis(ingresos_df, gastos_df)
        
        # Crear gráficas
        print("Generando gráficas...")
        charts = create_charts(merged, gastos_cat)
        
        # Crear reporte Excel
        print("Creando reporte Excel...")
        create_excel_report(merged, gastos_cat, ingresos_cat, charts)
        
        print(f"Reporte generado exitosamente: {REPORT_FILE}")
        
    except Exception as e:
        print(f"Error durante el análisis: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        if 'conn' in locals():
            conn.close()
        
        # Limpiar imágenes temporales
        for filename in os.listdir(IMAGE_FOLDER):
            file_path = os.path.join(IMAGE_FOLDER, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Error al eliminar {file_path}: {e}")

if __name__ == "__main__":
    main()